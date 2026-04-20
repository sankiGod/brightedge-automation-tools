# ============================================================
#  tools/parser.py — File parsing and ticket body parsing
#
#  Functions:
#    parse_brightedge_fields(text) → (username, mappings)
#    parse_file(file_bytes, filename) → { name: [rows] }
#    parse_csv(file_bytes)          → [rows]
#    parse_excel(file_bytes)        → { sheet_name: [rows] }
# ============================================================

import io
import re
import csv
import difflib
import openpyxl
import os

FUZZY_SHEET_THRESHOLD = float(os.environ.get("FUZZY_SHEET_THRESHOLD", "0.6"))
FUZZY_FILE_THRESHOLD  = float(os.environ.get("FUZZY_FILE_THRESHOLD", "0.5"))


# ─────────────────────────────────────────────
#  Ticket body parsing
# ─────────────────────────────────────────────

def parse_brightedge_fields(text):
    """
    Extracts BrightEdge username and identifier→account mappings from
    the ticket body. Parsing is flexible — labels are case-insensitive
    and extra surrounding text is ignored.

    Supported formats:

    Single account (CSV or single-sheet Excel):
        BrightEdge Username: john@company.com
        Account ID: 191

    Multiple CSVs:
        BrightEdge Username: john@company.com
        File: norgren_keywords.csv
        Account ID: 191
        File: bergfreunde_keywords.csv
        Account ID: 225

    Multi-sheet Excel:
        BrightEdge Username: john@company.com
        Sheet: Norgren
        Account ID: 191
        Sheet: Bergfreunde
        Account ID: 225

    Returns:
        (username, mappings) where:
          - username  is a string or None
          - mappings  is a list of {"identifier": str|None, "account_id": str}
                      or None if no Account IDs were found
    """
    # ── Username ─────────────────────────────────────────────
    username = None
    m = re.search(
        r"(?:brightedge\s+|be\s+)?user(?:name)?\s*:\s*(\S+@\S+\.\S+)",
        text,
        re.IGNORECASE,
    )
    if m:
        username = m.group(1).strip()

    # ── Identifier → Account ID mappings ────────────────────
    IDENTIFIER_RE = re.compile(
        r"(?:sheet|tab|worksheet|file(?:name)?)\s*[:#]?\s*(.+)",
        re.IGNORECASE,
    )
    ACCOUNT_RE = re.compile(
        r"(?:brightedge\s+|be\s+)?account(?:\s+id)?\s*[:#]?\s*(\d+)",
        re.IGNORECASE,
    )

    mappings           = []
    current_identifier = None

    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue

        id_match = IDENTIFIER_RE.match(stripped)
        if id_match:
            current_identifier = id_match.group(1).strip()
            continue

        acct_match = ACCOUNT_RE.match(stripped)
        if acct_match:
            mappings.append({
                "identifier": current_identifier,
                "account_id": acct_match.group(1).strip(),
            })
            current_identifier = None

    return username, (mappings if mappings else None)


# ─────────────────────────────────────────────
#  Column header normalisation + fuzzy matching
# ─────────────────────────────────────────────

# Known aliases for each standard field (after normalisation)
_KEYWORD_ALIASES = {"keyword", "kw", "keywords", "search term", "search terms"}
_PLP_ALIASES     = {"preferred landing page", "plp", "landing page", "url"}

def _normalize_header(h: str) -> str:
    """
    Normalises a column header for alias matching.
    Strips annotation suffixes, replaces whitespace chars, collapses spaces, lowercases.
    """
    h = re.sub(r"\s*[\(\[]?(optional|required)[\)\]]?\s*", "", h, flags=re.IGNORECASE)
    h = re.sub(r"[*#\[\]]", "", h)
    h = h.replace("\n", " ").replace("\t", " ")
    h = re.sub(r"\s+", " ", h).strip().lower()
    return h


def fuzzy_match_columns(headers: list) -> dict:
    """
    Rule-based column mapper using known aliases from CLAUDE.md.

    Normalises each header and checks against known aliases for keyword,
    plp, and keyword group columns. Falls back to difflib fuzzy matching
    if an exact alias is not found.

    Returns:
        {
            "keyword":    "<exact header name or None>",
            "plp":        "<exact header name or None>",
            "groups":     ["<col1>", ...],
            "confidence": float,   # 1.0 if keyword found, 0.0 if not
            "source":     "fuzzy",
        }

    confidence >= FUZZY_SHEET_THRESHOLD means the primary path was successful.
    confidence < threshold means the Claude fallback should be tried.
    """
    # Build normalised → original mapping
    norm_map = {_normalize_header(h): h for h in headers if h}

    keyword_col = None
    plp_col     = None
    group_cols  = []

    for norm, original in norm_map.items():
        if norm in _KEYWORD_ALIASES:
            keyword_col = original
        elif norm in _PLP_ALIASES:
            plp_col = original
        elif norm.startswith("keyword group"):
            group_cols.append(original)

    # Fuzzy fallback for keyword if not found via exact alias
    if keyword_col is None:
        close = difflib.get_close_matches(
            "keyword", list(norm_map.keys()), n=1, cutoff=FUZZY_SHEET_THRESHOLD
        )
        if close:
            keyword_col = norm_map[close[0]]
            print(f"  [Parser] Fuzzy matched keyword col: '{keyword_col}'")

    # Fuzzy fallback for plp
    if plp_col is None:
        for alias in ("preferred landing page", "plp", "landing page", "url"):
            close = difflib.get_close_matches(
                alias, list(norm_map.keys()), n=1, cutoff=FUZZY_SHEET_THRESHOLD
            )
            if close:
                plp_col = norm_map[close[0]]
                print(f"  [Parser] Fuzzy matched plp col: '{plp_col}'")
                break

    # Sort group columns by their original order in headers
    group_cols = [h for h in headers if h in group_cols]

    confidence = 1.0 if keyword_col else 0.0

    return {
        "keyword":    keyword_col,
        "plp":        plp_col,
        "groups":     group_cols,
        "confidence": confidence,
        "source":     "fuzzy",
    }


def _fuzzy_match(name, available, threshold):
    """
    Core fuzzy matcher — finds the closest match for `name` in `available`.
    Case-insensitive. Returns the matched name (original casing) or None.
    """
    name_lower      = name.lower().strip()
    available_lower = {s.lower().strip(): s for s in available}

    if name_lower in available_lower:
        return available_lower[name_lower]

    close = difflib.get_close_matches(
        name_lower, available_lower.keys(), n=1, cutoff=threshold
    )
    if close:
        matched = available_lower[close[0]]
        print(f"  Fuzzy matched '{name}' → '{matched}'")
        return matched

    return None


def fuzzy_match_sheet(name, available_sheets):
    """Matches an Excel sheet label against available tab names (threshold=0.6)."""
    return _fuzzy_match(name, available_sheets, FUZZY_SHEET_THRESHOLD)


def fuzzy_match_filename(name, available_files):
    """
    Matches a File: label against actual filenames (threshold=0.5).

    Tries two passes:
    1. Full filename match — "test_csv.csv" vs "test_csv.csv"
    2. Basename-only match — strips extension from both sides before comparing,
       so "test.csv" matches "test_csv.csv" (basenames: "test" vs "test_csv")

    Lower threshold handles short labels like 'norgren' matching
    'norgren_keywords.csv'.
    """
    # Pass 1: full filename
    result = _fuzzy_match(name, available_files, FUZZY_FILE_THRESHOLD)
    if result:
        return result

    # Pass 2: basename-only (strip extension from identifier and candidates)
    name_base       = name.rsplit(".", 1)[0]
    base_to_full    = {f.rsplit(".", 1)[0]: f for f in available_files}
    result_base     = _fuzzy_match(name_base, list(base_to_full.keys()), FUZZY_FILE_THRESHOLD)
    if result_base:
        matched = base_to_full[result_base]
        print(f"  Fuzzy matched '{name}' → '{matched}' (basename match)")
        return matched

    return None


# ─────────────────────────────────────────────
#  File parsing
# ─────────────────────────────────────────────

def parse_file(file_bytes, filename):
    """
    Routes to CSV or Excel parser based on file extension.

    Returns:
        { name: [rows] }
        For CSV:   { "__csv__": [rows] }
        For Excel: { "Sheet1": [rows], "Sheet2": [rows], ... }
    """
    import os
    _, ext = os.path.splitext(filename.lower())
    if ext in (".xlsx", ".xls"):
        print(f"  Detected Excel file: {filename}")
        return parse_excel(file_bytes)
    else:
        print(f"  Detected CSV file: {filename}")
        rows = parse_csv(file_bytes)
        return {"__csv__": rows} if rows else {}


def parse_csv(file_bytes):
    """
    Parses a CSV or TSV file into a list of normalised row dicts.
    Handles UTF-8 BOM that Excel-exported CSVs often include.
    Auto-detects tab delimiter: if the first line contains tabs but no commas,
    treats the file as TSV.
    All column names are lowercased and stripped.

    Returns: [{ col: value, ... }, ...]
    """
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")

    # Auto-detect delimiter from first non-empty line
    first_line = next((l for l in text.splitlines() if l.strip()), "")
    delimiter  = "\t" if "\t" in first_line and "," not in first_line else ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows   = []
    for row in reader:
        normalised = {k.strip().lower(): v.strip() for k, v in row.items() if k}
        rows.append(normalised)

    fmt = "TSV" if delimiter == "\t" else "CSV"
    if rows:
        print(f"  {fmt}: {len(rows)} rows — columns: {list(rows[0].keys())}")
    else:
        print(f"  {fmt}: no rows found.")

    return rows


def parse_excel(file_bytes):
    """
    Parses ALL sheets in an Excel file.
    Skips instruction rows at the top by finding the first row
    that contains a cell exactly matching 'keyword'.

    Returns: { sheet_name: [rows], ... }
    Only includes sheets that contain keyword data.
    """
    wb     = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}

    print(f"  Excel sheets found: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        rows = _parse_sheet(wb[sheet_name])
        if rows:
            result[sheet_name] = rows
            print(f"  '{sheet_name}': {len(rows)} rows")
        else:
            print(f"  '{sheet_name}': no keyword data — skipped")

    return result


def _parse_sheet(ws):
    """
    Parses a single Excel worksheet.
    Finds the header row by looking for a cell containing 'keyword',
    normalises column names, and returns all data rows below the header.
    """
    all_rows   = list(ws.iter_rows(values_only=True))
    header_idx = None

    for i, row in enumerate(all_rows):
        if any(str(c).strip().lower() == "keyword" for c in row if c):
            header_idx = i
            break

    if header_idx is None:
        return []

    headers = [
        str(h).strip().lower().replace("\n", " ").strip() if h else ""
        for h in all_rows[header_idx]
    ]

    rows = []
    for raw_row in all_rows[header_idx + 1:]:
        row_dict = {
            headers[i]: str(v).strip() if v is not None else ""
            for i, v in enumerate(raw_row)
            if i < len(headers) and headers[i]
        }
        if not any(row_dict.values()):
            continue
        kw = row_dict.get("keyword", "")
        if not kw or kw.isdigit():
            continue
        rows.append(row_dict)

    return rows


# ─────────────────────────────────────────────
#  KWG SE Upload — column mapping and Excel parsing
# ─────────────────────────────────────────────

_KWG_NAME_ALIASES = {
    "kwg name", "kwg", "keyword group name", "keyword group",
    "group name", "group",
}
_SE_IDS_ALIASES = {
    "search engine id list", "search engine ids", "se id list", "se ids",
    "search engine id", "se id", "se list", "se", "ses", "search engines",
}


def map_kwg_se_columns(headers: list) -> dict:
    """
    Maps raw column headers to the two standard KWG SE field names.

    Returns:
        {"kwg_name": "<col name or None>", "se_ids": "<col name or None>"}

    Raises:
        ValueError if neither KWG name nor SE ID column can be identified.
    """
    norm_map     = {_normalize_header(h): h for h in headers if h}
    kwg_name_col = None
    se_ids_col   = None

    for norm, original in norm_map.items():
        if norm in _KWG_NAME_ALIASES:
            kwg_name_col = original
        elif norm in _SE_IDS_ALIASES:
            se_ids_col = original

    # Fuzzy fallback for KWG name
    if kwg_name_col is None:
        for alias in ("kwg name", "keyword group name", "group name"):
            close = difflib.get_close_matches(
                alias, list(norm_map.keys()), n=1, cutoff=FUZZY_SHEET_THRESHOLD
            )
            if close:
                kwg_name_col = norm_map[close[0]]
                print(f"  [Parser] Fuzzy matched KWG name col: '{kwg_name_col}'")
                break

    # Fuzzy fallback for SE IDs
    if se_ids_col is None:
        for alias in ("search engine id list", "se ids", "se id"):
            close = difflib.get_close_matches(
                alias, list(norm_map.keys()), n=1, cutoff=FUZZY_SHEET_THRESHOLD
            )
            if close:
                se_ids_col = norm_map[close[0]]
                print(f"  [Parser] Fuzzy matched SE ID col: '{se_ids_col}'")
                break

    if not kwg_name_col and not se_ids_col:
        raise ValueError(
            f"Could not identify KWG name or SE ID columns. "
            f"Headers seen: {headers}"
        )

    return {"kwg_name": kwg_name_col, "se_ids": se_ids_col}


def parse_excel_kwg_se(file_bytes):
    """
    Parses all sheets from an Excel file for KWG SE data.
    Like parse_excel() but finds the header row by looking for KWG-related
    column names instead of 'keyword'.

    Returns: { sheet_name: [rows], ... }
    """
    wb     = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}

    print(f"  Excel sheets found: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        rows = _parse_sheet_kwg_se(wb[sheet_name])
        if rows:
            result[sheet_name] = rows
            print(f"  '{sheet_name}': {len(rows)} rows")
        else:
            print(f"  '{sheet_name}': no KWG SE data — skipped")

    return result


def _parse_sheet_kwg_se(ws):
    """
    Parses a single worksheet for KWG SE data.
    Finds the header row by looking for any KWG-indicator cell value.
    """
    all_rows        = list(ws.iter_rows(values_only=True))
    header_idx      = None
    kwg_indicators  = {"kwg name", "keyword group name", "group name", "kwg", "group"}

    for i, row in enumerate(all_rows):
        row_norms = {str(c).strip().lower() for c in row if c}
        if row_norms & kwg_indicators:
            header_idx = i
            break

    if header_idx is None:
        return []

    headers = [
        str(h).strip().lower().replace("\n", " ").strip() if h else ""
        for h in all_rows[header_idx]
    ]

    rows = []
    for raw_row in all_rows[header_idx + 1:]:
        row_dict = {
            headers[i]: str(v).strip() if v is not None else ""
            for i, v in enumerate(raw_row)
            if i < len(headers) and headers[i]
        }
        if not any(row_dict.values()):
            continue
        rows.append(row_dict)

    return rows
